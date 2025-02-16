from __future__ import unicode_literals
import frappe
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook
import re
from frappe import _
import frappe
from frappe.model.document import Document
from datetime import date, timedelta, datetime,time
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
	nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime,today, format_date)
import math
from frappe.utils import add_months, cint, flt, getdate, time_diff_in_hours,time_diff_in_seconds
import locale
import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
import openpyxl.styles as styles

@frappe.whitelist()
def download(start_date,end_date):
	filename = 'Today Canteen Count Report'
	max_in = datetime.strptime('06:30', '%H:%M').time()
	current_time = datetime.now().time()
	if current_time > max_in :
		args = {'start_date':today(),'end_date':today()}
	else:
		args = {'start_date':add_days(today(),-1),'end_date':add_days(today(),-1)}
	frappe.msgprint("Report is generating in the background,kindly check after few mins in the same page.")
	enqueue(today_canteen_count_report, queue='default', timeout=6000, event='today_canteen_count_report',filename=filename,args=args)
	
def make_xlsx(data,args, sheet_name=None, wb=None, column_widths=None):
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook()
		 
	ws = wb.create_sheet(sheet_name, 0)
	header_date = get_title()
	ws.append(header_date)

	header_date = get_title2(args)
	ws.append(header_date)

	header_date = get_title1()
	ws.append(header_date)

	header_date = get_title3()
	ws.append(header_date)
	
	data=get_data(args)
	for d in data:
		ws.append(d)
	
	data=get_data1(args)
	for d in data:
		ws.append(d)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column= len(get_title3()) )
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column= len(get_title3()) )
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column= 2 )
	ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column= 2 + ((len(get_title3()) - 2)/3) )
	ws.merge_cells(start_row=3, start_column=3 + ((len(get_title3()) - 2)/3), end_row=3, end_column= 2 + ((len(get_title3()) - 2)/3) + ((len(get_title3()) - 2)/3) )
	ws.merge_cells(start_row=3, start_column=3 + ((len(get_title3()) - 2)/3) + ((len(get_title3()) - 2)/3), end_row=3, end_column= len(get_title3()) )
	ws.merge_cells(start_row=len(get_data(args)) + 5, start_column=1, end_row=len(get_data(args)) + 5, end_column= 2)
	
	align_center = Alignment(horizontal='center',vertical='center')
	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin'))
	for rows in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=len(get_title3())):
		for cell in rows:
			cell.font = Font(bold=True)
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=3, min_col=3, max_row=len(get_data(args)) + 5, max_col= len(get_title3()) ): 
		for cell in rows:
			cell.fill = PatternFill(fgColor='A39EA0', fill_type="solid")
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=3, min_col=3 + ((len(get_title3()) - 2) // 3), max_row=len(get_data(args)) + 5, max_col=2 + ((len(get_title3()) - 2) // 3) + ((len(get_title3()) - 2) // 3)):
		for cell in rows:
			cell.fill = PatternFill(fgColor='FBF9FA', fill_type="solid")
			cell.alignment = align_center
	for rows in ws.iter_rows(min_row=1, min_col=1, max_row=len(get_data(args)) + 5, max_col= len(get_title3()) ):
		for cell in rows:
			cell.border = border

	ws.column_dimensions['A'].width = 20
	ws.column_dimensions['B'].width = 30

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def today_canteen_count_report(filename,args):
	xlsx_file = make_xlsx(filename,args)
	ret = frappe.get_doc({
			"doctype": "File",
			"attached_to_name": ' ',
			"attached_to_doctype": 'Report Dashboard',
			"attached_to_field": 'attach',
			"file_name": filename + '.xlsx',
			"is_private": 0,
			"content": xlsx_file.getvalue(),
			"decode": False
		})
	ret.save(ignore_permissions=True)
	frappe.db.commit()
	attached_file = frappe.get_doc("File", ret.name)
	frappe.db.set_value('Report Dashboard',None,'attach',attached_file.file_url)

@frappe.whitelist()
def get_title():
	status = []
	status = ['DongWoo Surfacetech (India) Pvt Ltd']
	return status

@frappe.whitelist()
def get_title2(args):
	status = []
	status = ['Today Canteen Count Summary ' + format_date((args['start_date']))]
	return status

@frappe.whitelist()
def get_title1():
	row = []
	row += ["Shift"," "]
	shift_types = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift_type in shift_types:
		count = frappe.db.count("Contractor")
		row += [shift_type.name]
		for _ in range(count +	4):
			row += [" "]
	return row

@frappe.whitelist()
def get_title3():
	result = []
	result += ["Parent","Department"]
	shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift in shifts:
		ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
		for e in ec:
			if e.name == "Staff":
				result += ["Staff"]
			elif e.name == "Worker":
				result += ["Wrks"]
			elif e.name == "D . Trainee":
				result += ["D . Trainee"]
			elif e.name == "NAPS":
				result += ["NAPS"]
			elif e.name == "Contract Employee":
				contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
				for contractor in contractors:
					result += [contractor.name]
		result += ["Total"]
	return result

@frappe.whitelist()
def get_data(args):
	status = []
	departments = frappe.db.sql("""
		SELECT * FROM `tabDepartment` 
		WHERE name != "All Departments" AND is_group = 1 
		ORDER BY `order_value` ASC
	""", as_dict=True)
	for department in departments:
		sub_departments = frappe.db.get_all("Department", filters={'parent_department': department.name}, fields=['*'])
		for sub_department in sub_departments:
			row = [department.name, sub_department.name]
			shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
			for shift in shifts:
				tot = 0
				ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
				for employee_type in ec:
					if employee_type.name != "Contract Employee":
						c = frappe.db.count("Attendance", {
							'attendance_date': (args['start_date']),
							'docstatus': ('!=', '2'),
							'shift': shift.name,
							'employee_type': employee_type.name,
							'in_time': ('!=', ''),
							'department': sub_department.name
						})
						if shift.name == "A":
							b_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "B",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'in_time': ('!=', ''),
								'out_time': ('=', '')
							})
							c_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "C",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'in_time': ('!=', ''),
								'out_time': ('=', '')
							})
							ot = b_shift_ot + c_shift_ot
						if shift.name == "B":
							c_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),-1),
								'docstatus': ('!=', '2'),
								'shift': "C",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'in_time': ('!=', ''),
								'out_time': ('=', '')
							})
							a_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "A",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'in_time': ('!=', ''),
								'out_time': ('=', '')
							})
							ot = a_shift_ot + c_shift_ot
						else:
							a_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "A",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'in_time': ('!=', ''),
								'out_time': ('=', '')
							})
							b_shift_ot = frappe.db.count("Attendance", {
								'attendance_date': add_days((args['start_date']),0),
								'docstatus': ('!=', '2'),
								'shift': "B",
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'in_time': ('!=', ''),
								'out_time': ('=', '')
							})
							ot = a_shift_ot + b_shift_ot
						tot += c
						tot += ot
						if ot > 0 :
							row.append(str(c) + " + " + str(ot))
						else:
							row.append(str(c))
					else:
						contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
						for contractor in contractors:
							c = frappe.db.count("Attendance", {
								'attendance_date': (args['start_date']),
								'docstatus': ('!=', '2'),
								'shift': shift.name,
								'employee_type': employee_type.name,
								'department': sub_department.name,
								'in_time': ('!=', ''),
								'contractor': contractor.name
							})
							if shift.name == "A":
								b_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "B",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'in_time': ('!=', ''),
									'out_time': ('=', '')
								})
								c_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "C",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'in_time': ('!=', ''),
									'out_time': ('=', '')
								})
								ot = b_shift_ot + c_shift_ot
							if shift.name == "B":
								c_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),-1),
									'docstatus': ('!=', '2'),
									'shift': "C",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'in_time': ('!=', ''),
									'out_time': ('=', '')
								})
								a_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "A",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'in_time': ('!=', ''),
									'out_time': ('=', '')
								})
								ot = a_shift_ot + c_shift_ot
							else:
								a_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "A",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'in_time': ('!=', ''),
									'out_time': ('=', '')
								})
								b_shift_ot = frappe.db.count("Attendance", {
									'attendance_date': add_days((args['start_date']),0),
									'docstatus': ('!=', '2'),
									'shift': "B",
									'employee_type': employee_type.name,
									'department': sub_department.name,
									'contractor': contractor.name,
									'in_time': ('!=', ''),
									'out_time': ('=', '')
								})
								ot = a_shift_ot + b_shift_ot
							tot += c
							tot += ot
							if ot > 0 :
								row.append(str(c) + " + " + str(ot))
							else:
								row.append(str(c))		
				row.append(str(tot))
			status.append(row)		
	return status

@frappe.whitelist()
def get_data1(args):
	status = []
	row = ["Total",'']
	shifts = frappe.db.sql("""SELECT * FROM `tabShift Type` ORDER BY `name` ASC""", as_dict=True)
	for shift in shifts:
		tot = 0
		ec = frappe.db.sql("""SELECT * FROM `tabEmployee Type` ORDER BY `order` ASC""", as_dict=True)
		for employee_type in ec:
			if employee_type.name != "Contract Employee":
				c = frappe.db.count("Attendance", {
					'attendance_date': (args['start_date']),
					'docstatus': ('!=', '2'),
					'shift': shift.name,
					'in_time': ('!=', ''),
					'employee_type': employee_type.name,
				})
				if shift.name == "A":
					b_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "B",
						'employee_type': employee_type.name,
						'in_time': ('!=', ''),
						'out_time': ('=', '')
					})
					c_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "C",
						'employee_type': employee_type.name,
						'in_time': ('!=', ''),
						'out_time': ('=', '')
					})
					ot = b_shift_ot + c_shift_ot
				if shift.name == "B":
					c_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),-1),
						'docstatus': ('!=', '2'),
						'shift': "C",
						'employee_type': employee_type.name,
						'in_time': ('!=', ''),
						'out_time': ('=', '')
					})
					a_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "A",
						'employee_type': employee_type.name,
						'in_time': ('!=', ''),
						'out_time': ('=', '')
					})
					ot = a_shift_ot + c_shift_ot
				else:
					a_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "A",
						'employee_type': employee_type.name,
						'in_time': ('!=', ''),
						'out_time': ('=', '')
					})
					b_shift_ot = frappe.db.count("Attendance", {
						'attendance_date': add_days((args['start_date']),0),
						'docstatus': ('!=', '2'),
						'shift': "B",
						'employee_type': employee_type.name,
						'in_time': ('!=', ''),
						'out_time': ('=', '')
					})
					ot = a_shift_ot + b_shift_ot
				tot += c
				tot += ot
				if ot > 0 :
					row.append(str(c) + " + " + str(ot))
				else:
					row.append(str(c))
			else:
				contractors = frappe.db.sql("""SELECT * FROM `tabContractor` ORDER BY `name` ASC""", as_dict=True)
				for contractor in contractors:
					c = frappe.db.count("Attendance", {
						'attendance_date': (args['start_date']),
						'docstatus': ('!=', '2'),
						'shift': shift.name,
						'employee_type': employee_type.name,
						'in_time': ('!=', ''),
						'contractor': contractor.name
					})
					if shift.name == "A":
						b_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "B",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'in_time': ('!=', ''),
							'out_time': ('=', '')
						})
						c_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "C",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'in_time': ('!=', ''),
							'out_time': ('=', '')
						})
						ot = b_shift_ot + c_shift_ot
					if shift.name == "B":
						c_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),-1),
							'docstatus': ('!=', '2'),
							'shift': "C",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'in_time': ('!=', ''),
							'out_time': ('=', '')
						})
						a_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "A",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'in_time': ('!=', ''),
							'out_time': ('=', '')
						})
						ot = a_shift_ot + c_shift_ot
					else:
						a_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "A",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'in_time': ('!=', ''),
							'out_time': ('=', '')
						})
						b_shift_ot = frappe.db.count("Attendance", {
							'attendance_date': add_days((args['start_date']),0),
							'docstatus': ('!=', '2'),
							'shift': "B",
							'employee_type': employee_type.name,
							'contractor': contractor.name,
							'in_time': ('!=', ''),
							'out_time': ('=', '')
						})
						ot = a_shift_ot + b_shift_ot
					tot += c
					tot += ot
					if ot > 0 :
						row.append(str(c) + " + " + str(ot))
					else:
						row.append(str(c))		
		row.append(str(tot))
	status.append(row)
			
	return status